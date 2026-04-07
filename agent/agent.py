"""
agent.py — JW RPA Agent v2
===========================
เพิ่ม endpoint สำหรับ AR:
  POST /import/in  — นำเข้าใบแจ้งหนี้ขาย (IN)
  POST /import/re  — นำเข้ารับชำระหนี้ (RE)
  GET  /ar/open    — ดึงรายการ IN ที่ยังค้างชำระ
"""

from license_checker import verify_license
import sys
if not verify_license(): sys.exit(1)

from flask import Flask, request, jsonify
from flask_cors import CORS
import dbf
import os
import json
import configparser
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
import logging
from import_in import register_routes as register_in_routes

app = Flask(__name__)
CORS(app)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================
# อ่าน config.ini — ถ้าไม่มีใช้ค่า default
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.ini")

def load_config():
    cfg = configparser.ConfigParser()
    cfg.read(CONFIG_FILE, encoding="utf-8")
    return cfg.get("agent", "dbf_path", fallback=r"Z:\Aulgor")

def save_config(dbf_path: str):
    cfg = configparser.ConfigParser()
    cfg["agent"] = {"dbf_path": dbf_path, "port": str(PORT)}
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        cfg.write(f)

DBF_FOLDER = load_config()
# ============================================================
ENCODING   = "cp874"
PORT       = 9999

HISTORY_FILE = os.path.join(os.path.dirname(__file__), "usage_history.json")
PAYMENT_FILE = os.path.join(os.path.dirname(__file__), "payment_notify.json")
MAX_HISTORY  = 100

# ── Supabase config (สำหรับ auto-register) ────────────────────
SUPABASE_URL      = "https://hklwbvpasiukjxvrrkxb.supabase.co"
SUPABASE_ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImhrbHdidnBhc2l1a2p4dnJya3hiIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzQ4MDQyOTcsImV4cCI6MjA5MDM4MDI5N30.EQGlDfMs6HH3PwQ__Oo-Kc8Lzf47gSABUe90ttQWSGg"

def auto_register():
    """ลงทะเบียนเครื่องใน Supabase ถ้ายังไม่มี — สถานะ 'pending'"""
    try:
        import requests as req
        import platform
        from license_checker import get_machine_id
        machine_id = get_machine_id()
        pc_name    = platform.node()
        os_ver     = platform.version()
        headers    = {
            "apikey":        SUPABASE_ANON_KEY,
            "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
            "Content-Type":  "application/json",
            "Prefer":        "resolution=ignore-duplicates",
        }
        # เช็คว่ามี record อยู่แล้วไหม
        chk = req.get(
            f"{SUPABASE_URL}/rest/v1/licenses?machine_id=eq.{machine_id}&select=id,is_active",
            headers=headers, timeout=8
        )
        if chk.status_code == 200 and chk.json():
            logger.info(f"[Register] เครื่องนี้ลงทะเบียนแล้ว")
            return
        # สร้าง record ใหม่ สถานะ pending (is_active=False)
        payload = {
            "machine_id":     machine_id,
            "customer_name":  pc_name,
            "plan":           "pending",
            "is_active":      False,
            "expire_date":    None,
            "pc_name":        pc_name,
            "os_version":     os_ver,
        }
        res = req.post(
            f"{SUPABASE_URL}/rest/v1/licenses",
            headers=headers, json=payload, timeout=8
        )
        if res.status_code in (200, 201):
            logger.info(f"[Register] ลงทะเบียนเครื่องใหม่สำเร็จ: {pc_name} ({machine_id})")
        else:
            logger.warning(f"[Register] ลงทะเบียนไม่สำเร็จ: {res.text}")
    except Exception as e:
        logger.warning(f"[Register] error: {e}")   # เก็บสูงสุด 100 รายการ

def log_usage(module: str, count: int, status: str, dbf_path: str = None, room: str = None):
    """บันทึกประวัติการนำเข้าลงไฟล์ JSON"""
    try:
        history = []
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                history = json.load(f)
        path = dbf_path or DBF_FOLDER
        # ชื่อห้องจาก path เช่น Z:\Aulgor → Aulgor
        if not room:
            room = os.path.basename(path.rstrip("\\/")) or path
        entry = {
            "id": int(datetime.now().timestamp() * 1000),
            "module": module,
            "count": count,
            "room": room,
            "dbf_path": path,
            "status": status,
            "created_at": datetime.now().isoformat(),
        }
        history.insert(0, entry)
        history = history[:MAX_HISTORY]
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"log_usage error: {e}")

JNLTYP_MAP = {
    "JV": "05", "PV": "01", "RV": "02",
    "SV": "03", "UV": "??", "BW": "00",
}
BANK_ACCT_MAP = {
    "S1": "1113-01", "F1": "1114-01", "C1": "1112-01",
}
DEBIT_ACCT_BW = "1111-00"
USERID = "BIT9"
AR_ACCT  = "1130-01"

# ── GL Config (อ่านจาก GL_config.xlsx) ───────────────────────
GL_CONFIG_PATH = "E:\\ทดสอบระบบ RPA\\RE_template.xlsx"

def load_gl_config():
    """โหลด GL account mapping จาก GL_config.xlsx"""
    cfg = {
        "default": {"AR_DEFAULT":"1130-01","WHT":"1151-02","FEE":"5360-04","SUSPEND":"9999-99"},
        "ar": {},    # CUSCOD → ACCNUM
        "bank": {},  # BNKCOD → ACCNUM
    }
    try:
        import openpyxl
        wb = openpyxl.load_workbook(GL_CONFIG_PATH, data_only=True)

        # GL_Default sheet
        if "GL_บัญชีหลัก" in wb.sheetnames:
            ws = wb["GL_บัญชีหลัก"]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] and row[1]:
                    cfg["default"][str(row[0]).strip()] = str(row[1]).strip()

        # AR_Account sheet
        if "GL_ลูกหนี้ตามลูกค้า" in wb.sheetnames:
            ws = wb["GL_ลูกหนี้ตามลูกค้า"]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] and row[1]:
                    cfg["ar"][str(row[0]).strip()] = str(row[1]).strip()

        # Bank_Account sheet
        if "GL_บัญชีธนาคาร" in wb.sheetnames:
            ws = wb["GL_บัญชีธนาคาร"]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0] and row[1]:
                    cfg["bank"][str(row[0]).strip()] = str(row[1]).strip()

        logger.info(f"[Config] โหลด GL_config.xlsx สำเร็จ")
    except Exception as e:
        logger.warning(f"[Config] ใช้ค่า default: {e}")
    return cfg

def load_cust_names():
    """โหลดชื่อลูกค้าจาก ARCUST.DBF"""
    names = {}
    path = os.path.join(DBF_FOLDER, "ARMAS.DBF")
    try:
        t = dbf.Table(path, codepage=ENCODING)
        t.open(mode=dbf.READ_ONLY)
        fname = list(t.field_names)
        name_field = next((f for f in ['CUSNAM','CUSNAME','CUSTNAM','CUSTNAME'] if f in fname), None)
        for rec in t:
            if dbf.is_deleted(rec): continue
            cod = str(rec.CUSCOD).strip()
            name = str(getattr(rec, name_field, "") or "").strip() if name_field else ""
            if cod: names[cod] = name
        t.close()
    except Exception as e:
        logger.warning(f"load_cust_names: {e}")
    return names

def get_ar_acct(cfg, cuscod):
    return cfg["ar"].get(cuscod, cfg["ar"].get("*", cfg["default"]["AR_DEFAULT"]))

def get_bank_acct(cfg, bnkcod):
    return cfg["bank"].get(bnkcod, cfg["bank"].get("*", "1113-00"))

def get_default(cfg, key):
    return cfg["default"].get(key, "9999-99")   # บัญชีลูกหนี้การค้า (Dr. ฝั่ง IN)
# ============================================================


# ── helpers ──────────────────────────────────────────────────

def stkcod_to_gl(stkcod: str) -> str:
    """แปลง STKCOD → GL account: '4100-02-09' → '4100-02'"""
    parts = stkcod.strip().split("-")
    if len(parts) >= 2:
        return f"{parts[0]}-{parts[1]}"
    return stkcod

def get_fields(filepath):
    t = dbf.Table(filepath, codepage=ENCODING)
    t.open(mode=dbf.READ_ONLY)
    fields = list(t.field_names)
    t.close()
    return fields

def mk(fields, m):
    return {f: m[f] for f in fields if f in m}

def sstr(val, maxlen):
    s = str(val) if val is not None else ""
    enc = s.encode(ENCODING, errors="replace")
    return enc[:maxlen].decode(ENCODING, errors="replace")

def parse_date(date_str):
    if not date_str:
        return date.today()
    if isinstance(date_str, date):
        return date_str
    s = str(date_str).strip()
    # DD/MM/YYYY หรือ DD/MM/YY
    if "/" in s:
        parts = s.split("/")
        day, month, yr = int(parts[0]), int(parts[1]), int(parts[2])
        if yr > 2400: yr -= 543
        elif yr < 100: yr = yr + 2500 - 543
        return date(yr, month, day)
    try:
        return datetime.fromisoformat(s).date()
    except:
        return date.today()

def calc_vat(total_incl, vatrat=7):
    net = Decimal(str(total_incl))
    rate = Decimal(str(vatrat))
    vatamt = (net * rate / (100 + rate)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    before = net - vatamt
    return float(before), float(vatamt), float(net)

def check_dup_artrn(docnum, rectyp=None):
    """ตรวจซ้ำใน ARTRN — ข้าม deleted records (ที่ลบผ่าน Express แล้ว)"""
    path = os.path.join(DBF_FOLDER, "ARTRN.DBF")
    try:
        t = dbf.Table(path, codepage=ENCODING)
        t.open(mode=dbf.READ_ONLY)
        for rec in t:
            if dbf.is_deleted(rec): continue   # ข้าม record ที่ลบแล้ว
            if str(rec.DOCNUM).strip() == docnum:
                if rectyp is None or str(rec.RECTYP).strip() == rectyp:
                    t.close(); return True
        t.close()
    except:
        pass
    return False

def get_in_remamt(docnum):
    path = os.path.join(DBF_FOLDER, "ARTRN.DBF")
    try:
        t = dbf.Table(path, codepage=ENCODING)
        t.open(mode=dbf.READ_ONLY)
        for rec in t:
            if str(rec.DOCNUM).strip() == docnum:
                result = float(rec.NETAMT or 0), float(rec.REMAMT or 0)
                t.close(); return result
        t.close()
    except:
        pass
    return 0.0, 0.0


# ── GL (เดิม) ─────────────────────────────────────────────────

def write_gl_records(pvs, jnltyp):
    hdr_path = os.path.join(DBF_FOLDER, "GLJNL.DBF")
    dtl_path = os.path.join(DBF_FOLDER, "GLJNLIT.DBF")
    hdr_fields = get_fields(hdr_path)
    dtl_fields = get_fields(dtl_path)
    th = dbf.Table(hdr_path, codepage=ENCODING)
    td = dbf.Table(dtl_path, codepage=ENCODING)
    th.open(mode=dbf.READ_WRITE)
    td.open(mode=dbf.READ_WRITE)
    today = date.today()
    ok = err = 0
    for pv in pvs:
        try:
            voucher = sstr(pv["docno"], 12)
            voudat  = parse_date(pv.get("date"))
            descrp  = sstr(pv.get("desc", ""), 50)
            hmap = {"JNLTYP":jnltyp,"BATCH":"","VOUDAT":voudat,"VOUCHER":voucher,"REFNUM":voucher,"SRCJNL":"GL","DESCRP":descrp,"REVERSE":" ","TRNSTAT":"P","DOCSTAT":"N","CREDAT":today,"CHGDAT":today}
            th.append(mk(hdr_fields, hmap))
            for seq, line in enumerate(pv.get("lines", []), start=1):
                acct  = sstr(line.get("acct", ""), 15)
                desc  = sstr(line.get("desc", descrp), 50)
                seqit = str(seq).zfill(2)
                debit  = float(line.get("debit",  0))
                credit = float(line.get("credit", 0))
                amount = debit if debit > 0 else credit
                trntyp = "0" if debit > 0 else "1"
                dmap = {"VOUCHER":voucher,"SEQIT":seqit,"VOUDAT":voudat,"ACCNUM":acct,"DESCRP":desc,"TRNTYP":trntyp,"AMOUNT":amount,"CHGDAT":today}
                td.append(mk(dtl_fields, dmap))
            ok += 1
        except Exception as e:
            logger.error(f"  ERR {pv.get('docno')}: {e}")
            err += 1
    th.close(); td.close()
    return ok, err

def write_bw_records(bws):
    hdr_path   = os.path.join(DBF_FOLDER, "GLJNL.DBF")
    dtl_path   = os.path.join(DBF_FOLDER, "GLJNLIT.DBF")
    bktrn_path = os.path.join(DBF_FOLDER, "BKTRN.DBF")
    hdr_fields   = get_fields(hdr_path)
    dtl_fields   = get_fields(dtl_path)
    bktrn_fields = get_fields(bktrn_path)
    th = dbf.Table(hdr_path, codepage=ENCODING); th.open(mode=dbf.READ_WRITE)
    td = dbf.Table(dtl_path, codepage=ENCODING); td.open(mode=dbf.READ_WRITE)
    tb = dbf.Table(bktrn_path, codepage=ENCODING); tb.open(mode=dbf.READ_WRITE)
    today = date.today()
    ok = err = 0
    for bw in bws:
        try:
            voucher = sstr(bw["docno"], 12)
            voudat  = parse_date(bw.get("date"))
            descrp  = sstr(bw.get("desc", ""), 50)
            amount  = float(bw.get("amount", 0))
            bank    = sstr(bw.get("bank", "S1"), 2)
            cr_acct = bw.get("cr_acct") or BANK_ACCT_MAP.get(bank, "1113-01")
            hmap = {"JNLTYP":"00","BATCH":"","VOUDAT":voudat,"VOUCHER":voucher,"REFNUM":voucher,"SRCJNL":"GL","DESCRP":descrp,"REVERSE":" ","TRNSTAT":"P","DOCSTAT":"N","CREDAT":today,"CHGDAT":today}
            th.append(mk(hdr_fields, hmap))
            td.append(mk(dtl_fields, {"VOUCHER":voucher,"SEQIT":"01","VOUDAT":voudat,"ACCNUM":sstr(DEBIT_ACCT_BW,15),"DESCRP":descrp,"TRNTYP":"0","AMOUNT":amount,"CHGDAT":today}))
            td.append(mk(dtl_fields, {"VOUCHER":voucher,"SEQIT":"02","VOUDAT":voudat,"ACCNUM":sstr(cr_acct,15),"DESCRP":descrp,"TRNTYP":"1","AMOUNT":amount,"CHGDAT":today}))
            bmap = {"BKTRNTYP":"BW","TRNDAT":voudat,"CHQNUM":voucher,"CHQDAT":voudat,"BNKACC":bank,"AMOUNT":amount,"NETAMT":amount,"VATAMT":0.0,"CHARGE":0.0,"REMAMT":0.0,"REMCUT":0.0,"JNLTRNTYP":"1","REMARK":descrp,"VOUCHER":voucher,"USERID":"RPA","CHGDAT":today}
            tb.append(mk(bktrn_fields, bmap))
            ok += 1
        except Exception as e:
            logger.error(f"  ERR BW {bw.get('docno')}: {e}")
            err += 1
    th.close(); td.close(); tb.close()
    return ok, err


# ── AR: IN (ใหม่) ────────────────────────────────────────────

def update_in_remamt(docnum, rcvamt):
    """หัก REMAMT ของ IN หลังรับชำระ และอัปเดต RCVAMT + DOCSTAT"""
    path = os.path.join(DBF_FOLDER, "ARTRN.DBF")
    try:
        t = dbf.Table(path, codepage=ENCODING)
        t.open(mode=dbf.READ_WRITE)
        for rec in t:
            if str(rec.DOCNUM).strip() == docnum and str(rec.RECTYP).strip() == "3":
                old_rem  = float(rec.REMAMT or 0)
                old_rcv  = float(rec.RCVAMT or 0)
                new_rem  = max(old_rem - rcvamt, 0.0)
                new_rcv  = old_rcv + rcvamt
                docstat  = "Y" if new_rem <= 0 else "N"
                with rec:
                    rec.REMAMT  = new_rem
                    rec.RCVAMT  = new_rcv
                    rec.DOCSTAT = docstat
                break
        t.close()
    except Exception as e:
        logger.error(f"update_in_remamt {docnum}: {e}")


def write_one_in(artrn_fields, stcrd_fields, gljnl_fields, gljnlit_fields,
                 th, ts, tgl, tdl, inv):
    """เขียน 1 IN record ลง ARTRN + STCRD + GLJNL + GLJNLIT"""
    docnum  = sstr(inv["docnum"], 12)
    docdat  = parse_date(inv.get("docdat"))
    cuscod  = sstr(inv.get("cuscod", ""), 10)
    youref   = sstr(inv.get("youref", ""), 30)
    custname = sstr(inv.get("custname", ""), 50)
    gl_desc  = sstr(f"ขายเชื่อให้  {custname}" if custname else f"ขายเชื่อให้  {cuscod}", 50)
    flgvat  = str(inv.get("flgvat", "2"))
    paytrm  = int(inv.get("paytrm", 0))
    vatrat  = 7
    items   = inv.get("items", [])
    total_incl = sum(float(it.get("trnval", 0)) for it in items)

    if flgvat == "2":          # รวม VAT → ถอย VAT ออก
        amount_bf, vatamt, netamt = calc_vat(total_incl, vatrat)
    elif flgvat == "1":        # แยก VAT → บวก VAT เพิ่ม
        amount_bf = total_incl
        vatamt = round(total_incl * vatrat / 100, 2)
        netamt = total_incl + vatamt
    else:                      # "0" = ไม่มี VAT
        amount_bf = total_incl
        vatamt = 0.0
        netamt = total_incl

    today = date.today()
    hmap = {
        "RECTYP":"3","DOCNUM":docnum,"DOCDAT":docdat,"POSTGL":" ",
        "SONUM":" "*12,"CNTYP":" ","DEPCOD":" "*4,"FLGVAT":flgvat,
        "SLMCOD":" "*10,"CUSCOD":cuscod,"SHIPTO":" "*10,"YOUREF":youref,
        "AREACOD":" "*4,"PAYTRM":paytrm,"DUEDAT":docdat,"BILNUM":"~",
        "NXTSEQ":" -1","AMOUNT":amount_bf,"DISC":" "*10,"DISCAMT":0.0,
        "AFTDISC":amount_bf,"ADVNUM":"AI","ADVAMT":0.0,"TOTAL":amount_bf,
        "AMTRAT0":0.0,"VATRAT":vatrat if flgvat != "0" else None,"VATAMT":vatamt,"NETAMT":netamt,
        "NETVAL":netamt,"RCVAMT":0.0,"REMAMT":netamt,"COMAMT":0.0,
        "CMPLAPP":"N","CMPLDAT":None,"DOCSTAT":"N","CSHRCV":0.0,
        "CHQRCV":0.0,"INTRCV":0.0,"BEFTAX":0.0,"TAXRAT":None,
        "TAXCOND":" ","TAX":0.0,"IVCAMT":0.0,"CHQPAS":0.0,
        "VATDAT":None,"VATPRD":None,"VATLATE":" ","SRV_VATTYP":"2",
        "DLVBY":"  ","RESERVE":None,"USERID":USERID,"CHGDAT":today,
        "USERPRN":" "*8,"PRNDAT":None,"PRNCNT":None,"PRNTIM":" "*8,
        "AUTHID":" "*8,"APPROVE":None,"BILLTO":" "*10,"ORGNUM":0,
    }
    th.append(mk(artrn_fields, hmap))

    for i, it in enumerate(items, 1):
        stkcod  = sstr(it.get("stkcod",""), 20)
        stkdes  = sstr(it.get("stkdes",""), 50)
        loccod  = sstr(it.get("loccod","01"), 4)
        trnqty  = float(it.get("trnqty", 1))
        tqucod  = sstr(it.get("tqucod","AA"), 2)
        unitpr  = float(it.get("unitpr", 0))
        discamt = float(it.get("discamt", 0))
        trnval  = float(it.get("trnval", trnqty*unitpr-discamt))
        dmap = {
            "STKCOD":stkcod,"LOCCOD":loccod,"DOCNUM":docnum,
            "SEQNUM":f"{i:>3}","DOCDAT":docdat,"RDOCNUM":" "*15,
            "REFNUM":" "*15,"DEPCOD":" "*4,"POSOPR":"9","FREE":" ",
            "VATCOD":" ","PEOPLE":cuscod,"SLMCOD":" "*10,"FLAG":" ",
            "TRNQTY":trnqty,"TQUCOD":tqucod,"TFACTOR":1.0,"UNITPR":unitpr,
            "DISC":" "*10,"DISCAMT":discamt,"TRNVAL":trnval,"PHYBAL":0.0,
            "RETSTK":" ","XTRNQTY":0.0,"XUNITPR":0.0,"XTRNVAL":0.0,
            "XSALVAL":0.0,"NETVAL":trnval,"MLOTNUM":" "*24,"MREMBAL":0.0,
            "MREMVAL":0.0,"BALCHG":0.0,"VALCHG":0.0,"LOTBAL":0.0,
            "LOTVAL":0.0,"LUNITPR":0.0,"PSTKCOD":" "*20,
            "ACCNUMDR":" "*15,"ACCNUMCR":" "*15,"STKDES":stkdes,
            "PACKING":" "*15,"JOBCOD":" "*6,"PHASE":" "*4,
            "COSCOD":" "*4,"REIMBURSE":" ",
        }
        ts.append(mk(stcrd_fields, dmap))

    # ── GLJNL header ──
    tgl.append(mk(gljnl_fields, {
        "JNLTYP":"03","BATCH":"","VOUDAT":docdat,"VOUCHER":docnum,
        "REFNUM":docnum,"SRCJNL":"GL","DESCRP":gl_desc,
        "REVERSE":" ","TRNSTAT":"P","DOCSTAT":"N",
        "CREDAT":today,"CHGDAT":today,
    }))

    # ── GLJNLIT lines ──
    # Dr. ลูกหนี้ (netamt รวม) — ใช้ stkdes ของรายการแรกเป็นคำอธิบาย
    first_desc = gl_desc
    tdl.append(mk(gljnlit_fields, {
        "VOUCHER":docnum,"SEQIT":"01","VOUDAT":docdat,
        "ACCNUM":sstr(AR_ACCT, 15),"DESCRP":first_desc,
        "TRNTYP":"0","AMOUNT":netamt,"CHGDAT":today,
    }))
    # Cr. รายได้/บัญชีพัก ตาม STKCOD — ใช้ stkdes เป็นคำอธิบาย
    for seq, it in enumerate(items, 2):
        acct   = stkcod_to_gl(it.get("stkcod", ""))
        trnval = float(it.get("trnval", 0))
        desc   = sstr(f"ขายเชื่อให้ {custname}".strip(), 50)
        tdl.append(mk(gljnlit_fields, {
            "VOUCHER":docnum,"SEQIT":str(seq).zfill(2),"VOUDAT":docdat,
            "ACCNUM":sstr(acct, 15),"DESCRP":desc,
            "TRNTYP":"1","AMOUNT":trnval,"CHGDAT":today,
        }))

    # ── GLBAL: อัปเดตยอดสะสม ──
    gl_entries = [(AR_ACCT, "0", netamt)]
    for it in items:
        gl_entries.append((stkcod_to_gl(it.get("stkcod","")), "1", float(it.get("trnval",0))))
    update_glbal(gl_entries, docdat)

    return netamt


# ── AR: RE (v3 synced) ────────────────────────────────────────

def write_one_re(artrn_fields, arrcpit_fields, arrcpcq_fields, bktrn_fields,
                 gljnl_fields, gljnlit_fields,
                 th, tp, tq, tb, tgl, tdl, rcp):
    """เขียน 1 RE record ลง ARTRN + ARRCPIT + ARRCPCQ + BKTRN + GLJNL + GLJNLIT (v3)"""
    rcpnum  = sstr(rcp["rcpnum"], 12)
    rcpdat  = parse_date(rcp.get("rcpdat"))
    cuscod  = sstr(rcp.get("cuscod",""), 10)
    bnkcod  = sstr(rcp.get("bnkcod",""), 2) or "  "
    paytyp  = str(rcp.get("paytyp","T")).upper()
    chqnum  = sstr(rcp.get("chqnum",""), 15)
    chqdat  = parse_date(rcp.get("chqdat")) if rcp.get("chqdat") else rcpdat
    remark  = sstr(rcp.get("remark",""), 50)
    whtamt  = float(rcp.get("whtamt",   0))
    fee     = float(rcp.get("fee",      0))
    suspend = float(rcp.get("suspend",  0))
    items   = rcp.get("items", [])
    total   = sum(float(it.get("rcvamt",0)) for it in items)
    # ใช้ transfer จาก frontend ถ้ามี ไม่งั้นคำนวณจาก total-wht-fee-suspend
    _tf     = float(rcp.get("transfer", 0))
    net_tf  = _tf if _tf > 0 else max(total - whtamt - fee - suspend, 0.0)
    today   = date.today()

    hmap = {
        "RECTYP"    :"9",
        "DOCNUM"    :rcpnum,
        "DOCDAT"    :rcpdat,
        "POSTGL"    :" ",
        "SONUM"     :" "*12,
        "CNTYP"     :" ",
        "DEPCOD"    :" "*4,
        "FLGVAT"    :"2",
        "SLMCOD"    :" "*10,
        "CUSCOD"    :cuscod,
        "SHIPTO"    :" "*10,
        "YOUREF"    :remark[:30],
        "AREACOD"   :" "*4,
        "PAYTRM"    :None,
        "DUEDAT"    :rcpdat,
        "BILNUM"    :"~",
        "NXTSEQ"    :"   ",
        "AMOUNT"    :0.0,
        "DISC"      :" "*10,
        "DISCAMT"   :0.0,
        "AFTDISC"   :0.0,
        "ADVNUM"    :" "*12,
        "ADVAMT"    :0.0,
        "TOTAL"     :0.0,
        "AMTRAT0"   :total,
        "VATRAT"    :None,
        "VATAMT"    :0.0,
        "NETAMT"    :total,
        "NETVAL"    :0.0,
        "RCVAMT"    :0.0,
        "REMAMT"    :0.0,
        "COMAMT"    :0.0,
        "CMPLAPP"   :"Y",
        "CMPLDAT"   :rcpdat,
        "DOCSTAT"   :"M",
        "CSHRCV"    :total if paytyp == "E" else 0.0,
        "CHQRCV"    :0.0   if paytyp == "E" else total,
        "INTRCV"    :0.0,
        "BEFTAX"    :0.0,
        "TAXRAT"    :None,
        "TAXCOND"   :" ",
        "TAX"       :0.0,
        "IVCAMT"    :total,
        "CHQPAS"    :total,
        "VATDAT"    :rcpdat,
        "VATPRD"    :None,
        "VATLATE"   :" ",
        "SRV_VATTYP":"-",
        "DLVBY"     :"  ",
        "RESERVE"   :None,
        "USERID"    :USERID,
        "CHGDAT"    :today,
        "USERPRN"   :" "*8,
        "PRNDAT"    :None,
        "PRNCNT"    :None,
        "PRNTIM"    :" "*8,
        "AUTHID"    :" "*8,
        "APPROVE"   :None,
        "BILLTO"    :" "*10,
        "ORGNUM"    :0,
    }
    th.append(mk(artrn_fields, hmap))

    trre = chqnum or f"TR{rcpnum}"
    txre = f"TX{rcpnum}"
    # T=โอน, C=เช็ก, E=สด(เงินสด) — เงินสดไม่มี BKTRN/ARRCPCQ
    is_cash = (paytyp == "E")

    # template bmap สำหรับ BKTRN (ใช้ร่วมกับ fee/suspend)
    bmap_tmpl = {
        "BKTRNTYP"  :"bR",   # Express ใช้ bR สำหรับโอน/เช็ค; เงินสดไม่มี BKTRN
        "TRNDAT"    :rcpdat,
        "CHQNUM"    :trre,
        "CHQDAT"    :chqdat,
        "BNKCOD"    :"  ",
        "BRANCH"    :" "*22,
        "CUSCOD"    :cuscod,
        "NAME"      :" "*50,
        "DEPCOD"    :" "*4,
        "POSTGL"    :" ",
        "GETDAT"    :rcpdat,
        "PAYINDAT"  :None,
        "AMOUNT"    :net_tf,
        "CHARGE"    :0.0,
        "VATAMT"    :0.0,
        "NETAMT"    :net_tf,
        "REMAMT"    :0.0,
        "REMCUT"    :0.0,
        "CMPLAPP"   :" ",
        "CHQSTAT"   :"19",
        "BNKACC"    :"  ",
        "JNLTRNTYP" :" ",
        "REMARK"    :remark[:50],
        "REFDOC"    :rcpnum,
        "REFNUM"    :trre[:15],
        "VATDAT"    :rcpdat,
        "VATPRD"    :date(rcpdat.year,rcpdat.month,1),
        "VATLATE"   :" ",
        "VATTYP"    :" ",
        "VOUCHER"   :" "*12,
        "USERID"    :USERID,
        "CHGDAT"    :today,
        "AUTHID"    :" "*8,
        "APPROVE"   :None,
        "TAXID"     :" "*15,
        "ORGNUM"    :0,
    }

    # BKTRN: ยอดโอนจริง (ข้ามถ้าเงินสด)
    if not is_cash:
        tb.append(mk(bktrn_fields, bmap_tmpl))

    # BKTRN: WHT (ถ้ามี)
    if whtamt > 0:
        bmap_wht = dict(bmap_tmpl)
        bmap_wht.update({"CHQNUM":txre,"AMOUNT":whtamt,"NETAMT":whtamt,"REFNUM":txre[:15]})
        tb.append(mk(bktrn_fields, bmap_wht))

    # หมายเหตุ: fee และ suspend ไม่มี BKTRN — Express ใช้เฉพาะ ARRCPCQ

    # ARRCPCQ — ยอดโอน (ข้ามถ้าเงินสด), WHT, fee (TFRE), suspend (99RE)
    if not is_cash:
        tq.append(mk(arrcpcq_fields, {"RCPNUM":rcpnum,"CHQNUM":trre,"RCVAMT":net_tf}))
    if whtamt > 0:
        tq.append(mk(arrcpcq_fields, {"RCPNUM":rcpnum,"CHQNUM":txre,"RCVAMT":whtamt}))
    if fee > 0:
        tq.append(mk(arrcpcq_fields, {"RCPNUM":rcpnum,"CHQNUM":f"TF{rcpnum}"[:12],"RCVAMT":fee}))
    if suspend > 0:
        tq.append(mk(arrcpcq_fields, {"RCPNUM":rcpnum,"CHQNUM":f"99{rcpnum}"[:12],"RCVAMT":suspend}))

    # ARRCPIT + หัก REMAMT ของ IN
    for it in items:
        docnum = sstr(it.get("docnum",""), 12)
        rcvamt = float(it.get("rcvamt", 0))
        vatamt = float(it.get("vatamt", 0))
        tp.append(mk(arrcpit_fields, {"RCPNUM":rcpnum,"DOCNUM":docnum,"RECTYP":"3","RCVAMT":rcvamt,"VATAMT":vatamt}))
        if docnum and rcvamt > 0:
            update_in_remamt(docnum, rcvamt)   # ตัดยอดลูกหนี้

    # ── GL + GLBAL สำหรับ RE ──
    cfg      = load_gl_config()
    ar_acc   = get_ar_acct(cfg, cuscod)
    bk_acc   = get_bank_acct(cfg, bnkcod)
    wht_acc  = get_default(cfg, "WHT")
    fee_acc  = get_default(cfg, "FEE")
    sus_acc  = get_default(cfg, "SUSPEND")
    # เงินสด: ใช้บัญชี CASH จาก config หรือ default 1111-00
    cash_acc = get_default(cfg, "CASH") if "CASH" in cfg["default"] else "1111-00"

    # Dr. บัญชีรับเงิน: เงินสด→cash_acc, โอน/เช็ค→bk_acc
    recv_acc = cash_acc if is_cash else bk_acc
    transfer = net_tf   # ยอดเงินรับจริง (โอนหรือเงินสด)
    custname = sstr(rcp.get("custname","") or cuscod, 30)
    gl_desc  = sstr(f"รับชำระหนี้  {custname}", 50)

    # Cr. ลูกหนี้ = จำนวนเงิน (total rcvamt) ← ยอดตามบิล
    cr_total = total   # ไม่ใช่ net_tf

    # GLJNL header
    if tgl is not None:
        tgl.append(mk(gljnl_fields, {
            "JNLTYP":"02","BATCH":"","VOUDAT":rcpdat,"VOUCHER":rcpnum,
            "REFNUM":rcpnum,"SRCJNL":"GL","DESCRP":gl_desc,
            "REVERSE":" ","TRNSTAT":"P","DOCSTAT":"N",
            "CREDAT":today,"CHGDAT":today,
        }))

        seq = 1
        gl_entries = []

        # Dr. รับเงิน (บัญชีธนาคาร หรือ เงินสด)
        if transfer > 0:
            tdl.append(mk(gljnlit_fields, {"VOUCHER":rcpnum,"SEQIT":str(seq).zfill(2),"VOUDAT":rcpdat,"ACCNUM":sstr(recv_acc,15),"DESCRP":gl_desc,"TRNTYP":"0","AMOUNT":transfer,"CHGDAT":today}))
            gl_entries.append((recv_acc,"0",transfer)); seq+=1

        # Dr. WHT
        if whtamt > 0:
            tdl.append(mk(gljnlit_fields, {"VOUCHER":rcpnum,"SEQIT":str(seq).zfill(2),"VOUDAT":rcpdat,"ACCNUM":sstr(wht_acc,15),"DESCRP":"ภาษีหัก ณ ที่จ่าย","TRNTYP":"0","AMOUNT":whtamt,"CHGDAT":today}))
            gl_entries.append((wht_acc,"0",whtamt)); seq+=1

        # Dr. ค่าธรรมเนียม
        if fee > 0:
            tdl.append(mk(gljnlit_fields, {"VOUCHER":rcpnum,"SEQIT":str(seq).zfill(2),"VOUDAT":rcpdat,"ACCNUM":sstr(fee_acc,15),"DESCRP":"ค่าธรรมเนียมธนาคาร","TRNTYP":"0","AMOUNT":fee,"CHGDAT":today}))
            gl_entries.append((fee_acc,"0",fee)); seq+=1

        # Dr. บัญชีพัก
        if suspend > 0:
            tdl.append(mk(gljnlit_fields, {"VOUCHER":rcpnum,"SEQIT":str(seq).zfill(2),"VOUDAT":rcpdat,"ACCNUM":sstr(sus_acc,15),"DESCRP":"บัญชีพัก","TRNTYP":"0","AMOUNT":suspend,"CHGDAT":today}))
            gl_entries.append((sus_acc,"0",suspend)); seq+=1

        # Cr. ลูกหนี้ = จำนวนเงินรวมตามบิล (transfer+wht+fee+suspend)
        tdl.append(mk(gljnlit_fields, {"VOUCHER":rcpnum,"SEQIT":str(seq).zfill(2),"VOUDAT":rcpdat,"ACCNUM":sstr(ar_acc,15),"DESCRP":gl_desc,"TRNTYP":"1","AMOUNT":cr_total,"CHGDAT":today}))
        gl_entries.append((ar_acc,"1",cr_total))

        # อัปเดต GLBAL
        update_glbal(gl_entries, rcpdat)


# ── GLBAL: อัปเดตยอดสะสมรายเดือน ─────────────────────────────

def update_glbal(gl_entries, docdat):
    """อัปเดต GLBAL.DBF ยอด DEBIT{m}/CREDIT{m} ตามเดือนของ docdat
    gl_entries = list of (accnum, trntyp, amount)
      trntyp '0' = debit, '1' = credit
    """
    glbal_path = os.path.join(DBF_FOLDER, "GLBAL.DBF")
    if not os.path.exists(glbal_path):
        logger.warning("GLBAL.DBF ไม่พบ — ข้าม")
        return
    try:
        month = docdat.month   # 1-12
        dr_field = f"DEBIT{month}"
        cr_field = f"CREDIT{month}"

        tb = dbf.Table(glbal_path, codepage=ENCODING)
        tb.open(mode=dbf.READ_WRITE)
        glbal_fields = list(tb.field_names)

        # รวม amount ต่อ account
        totals = {}   # accnum → (dr_total, cr_total)
        for accnum, trntyp, amount in gl_entries:
            accnum = accnum.strip()
            if accnum not in totals:
                totals[accnum] = [0.0, 0.0]
            if trntyp == "0":
                totals[accnum][0] += amount
            else:
                totals[accnum][1] += amount

        # หา record ที่มีอยู่
        found = {}
        for rec in tb:
            acc = str(rec.ACCNUM).strip()
            if acc in totals:
                found[acc] = rec

        for accnum, (dr, cr) in totals.items():
            if accnum in found:
                rec = found[accnum]
                with rec:
                    if dr_field in glbal_fields and dr > 0:
                        setattr(rec, dr_field, float(getattr(rec, dr_field) or 0) + dr)
                    if cr_field in glbal_fields and cr > 0:
                        setattr(rec, cr_field, float(getattr(rec, cr_field) or 0) + cr)
            else:
                # สร้าง record ใหม่
                new_rec = {f: 0.0 for f in glbal_fields if f.startswith(("DEBIT","CREDIT","BEG"))}
                new_rec["ACCNUM"] = accnum
                new_rec["DEPCOD"] = "    "
                new_rec["JOBCOD"] = "      "
                new_rec["CALSTA"] = " " * 24
                if dr > 0 and dr_field in glbal_fields:
                    new_rec[dr_field] = dr
                if cr > 0 and cr_field in glbal_fields:
                    new_rec[cr_field] = cr
                tb.append(mk(glbal_fields, new_rec))
            logger.info(f"  GLBAL {accnum}: DR{month}+={dr:.2f} CR{month}+={cr:.2f}")

        tb.close()
    except Exception as e:
        logger.error(f"update_glbal error: {e}")


# ── Routes ───────────────────────────────────────────────────

@app.route('/import', methods=['POST'])
def do_import():
    try:
        data  = request.json
        jtype = data.get("type","PV")
        rows  = data.get("data", [])
        logger.info(f"\n[Agent] นำเข้า {jtype} {len(rows)} รายการ")
        if jtype == "BW":
            ok, err = write_bw_records(rows)
        else:
            ok, err = write_gl_records(rows, JNLTYP_MAP.get(jtype,"05"))
        _jtype_names = {"JV":"ทั่วไป","PV":"จ่าย","RV":"รับ","SV":"ขาย","UV":"ซื้อ","BW":"ถอนเงิน"}
        _jtype_label = jtype + " — " + _jtype_names.get(jtype, jtype)
        log_usage(module=_jtype_label, count=ok, status="success" if err==0 else "error")
        return jsonify({"success":ok,"error":err})
    except Exception as e:
        return jsonify({"error":str(e)}), 500


@app.route('/import/in', methods=['POST'])
def do_import_in():
    """รับ JSON: { invoices: [ {docnum, docdat, cuscod, flgvat, paytrm, youref, items:[...]} ] }"""
    try:
        data     = request.json
        invoices = data.get("invoices", [])
        logger.info(f"\n[Agent] นำเข้า IN {len(invoices)} ใบ")

        artrn_path  = os.path.join(DBF_FOLDER, "ARTRN.DBF")
        stcrd_path  = os.path.join(DBF_FOLDER, "STCRD.DBF")
        gljnl_path  = os.path.join(DBF_FOLDER, "GLJNL.DBF")
        gljnlit_path= os.path.join(DBF_FOLDER, "GLJNLIT.DBF")
        artrn_f  = get_fields(artrn_path)
        stcrd_f  = get_fields(stcrd_path)
        gljnl_f  = get_fields(gljnl_path)
        gljnlit_f= get_fields(gljnlit_path)
        th  = dbf.Table(artrn_path,   codepage=ENCODING); th.open(mode=dbf.READ_WRITE)
        ts  = dbf.Table(stcrd_path,   codepage=ENCODING); ts.open(mode=dbf.READ_WRITE)
        tgl = dbf.Table(gljnl_path,   codepage=ENCODING); tgl.open(mode=dbf.READ_WRITE)
        tdl = dbf.Table(gljnlit_path, codepage=ENCODING); tdl.open(mode=dbf.READ_WRITE)

        ok = err = skip = 0
        details = []
        for inv in invoices:
            docnum = sstr(inv.get("docnum",""), 12)
            if check_dup_artrn(docnum):
                skip += 1
                details.append({"docnum":docnum,"status":"dup"})
                continue
            try:
                netamt = write_one_in(artrn_f, stcrd_f, gljnl_f, gljnlit_f, th, ts, tgl, tdl, inv)
                ok += 1
                details.append({"docnum":docnum,"status":"ok","netamt":netamt})
                logger.info(f"  OK {docnum}")
            except Exception as e:
                err += 1
                details.append({"docnum":docnum,"status":"err","msg":str(e)})
                logger.error(f"  ERR {docnum}: {e}")

        th.close(); ts.close(); tgl.close(); tdl.close()
        log_usage(module="INV — ใบแจ้งหนี้", count=ok, status="success" if err==0 else "error")
        return jsonify({"success":ok,"skipped":skip,"error":err,"details":details})
    except Exception as e:
        logger.error(f"import/in error: {e}")
        return jsonify({"error":str(e)}), 500


@app.route('/import/re', methods=['POST'])
def do_import_re():
    """รับ JSON: { receipts: [ {rcpnum, rcpdat, cuscod, bnkcod, chqnum, chqdat, whtamt, remark, items:[{docnum,rcvamt,vatamt}]} ] }"""
    try:
        data     = request.json
        receipts = data.get("receipts", [])
        logger.info(f"\n[Agent] นำเข้า RE {len(receipts)} ใบ")

        artrn_path   = os.path.join(DBF_FOLDER, "ARTRN.DBF")
        arrcpit_path = os.path.join(DBF_FOLDER, "ARRCPIT.DBF")
        arrcpcq_path = os.path.join(DBF_FOLDER, "ARRCPCQ.DBF")
        bktrn_path   = os.path.join(DBF_FOLDER, "BKTRN.DBF")
        artrn_f   = get_fields(artrn_path)
        arrcpit_f = get_fields(arrcpit_path)
        arrcpcq_f = get_fields(arrcpcq_path)
        bktrn_f   = get_fields(bktrn_path)

        gljnl_path   = os.path.join(DBF_FOLDER, "GLJNL.DBF")
        gljnlit_path = os.path.join(DBF_FOLDER, "GLJNLIT.DBF")
        gljnl_f      = get_fields(gljnl_path)
        gljnlit_f    = get_fields(gljnlit_path)

        th  = dbf.Table(artrn_path,   codepage=ENCODING); th.open(mode=dbf.READ_WRITE)
        tp  = dbf.Table(arrcpit_path, codepage=ENCODING); tp.open(mode=dbf.READ_WRITE)
        tq  = dbf.Table(arrcpcq_path, codepage=ENCODING); tq.open(mode=dbf.READ_WRITE)
        tb  = dbf.Table(bktrn_path,   codepage=ENCODING); tb.open(mode=dbf.READ_WRITE)
        tgl = dbf.Table(gljnl_path,   codepage=ENCODING); tgl.open(mode=dbf.READ_WRITE)
        tdl = dbf.Table(gljnlit_path, codepage=ENCODING); tdl.open(mode=dbf.READ_WRITE)

        ok = err = skip = 0
        details = []
        for rcp in receipts:
            rcpnum = sstr(rcp.get("rcpnum",""), 12)
            if check_dup_artrn(rcpnum, rectyp="9"):
                skip += 1
                details.append({"rcpnum":rcpnum,"status":"dup"})
                continue
            try:
                write_one_re(artrn_f, arrcpit_f, arrcpcq_f, bktrn_f, gljnl_f, gljnlit_f,
                             th, tp, tq, tb, tgl, tdl, rcp)
                ok += 1
                details.append({"rcpnum":rcpnum,"status":"ok"})
                logger.info(f"  OK {rcpnum}")
            except Exception as e:
                err += 1
                details.append({"rcpnum":rcpnum,"status":"err","msg":str(e)})
                logger.error(f"  ERR {rcpnum}: {e}")

        th.close(); tp.close(); tq.close(); tb.close(); tgl.close(); tdl.close()
        log_usage(module="RE — รับชำระหนี้", count=ok, status="success" if err==0 else "error")
        return jsonify({"success":ok,"skipped":skip,"error":err,"details":details})
    except Exception as e:
        logger.error(f"import/re error: {e}")
        return jsonify({"error":str(e)}), 500


@app.route('/ar/open', methods=['GET'])
def get_open_ar():
    """ดึงรายการ IN ที่ยังค้างชำระ (REMAMT > 0)"""
    try:
        cuscod_filter = request.args.get("cuscod","")
        artrn_path = os.path.join(DBF_FOLDER, "ARTRN.DBF")
        cust_names = load_cust_names()
        result = []
        t = dbf.Table(artrn_path, codepage=ENCODING)
        t.open(mode=dbf.READ_ONLY)
        for rec in t:
            if str(rec.RECTYP).strip() != "3": continue
            remamt = float(rec.REMAMT or 0)
            if remamt <= 0: continue
            cuscod = str(rec.CUSCOD).strip()
            if cuscod_filter and cuscod != cuscod_filter: continue
            result.append({
                "docnum"  : str(rec.DOCNUM).strip(),
                "docdat"  : str(rec.DOCDAT),
                "cuscod"  : cuscod,
                "cusname" : cust_names.get(cuscod, ""),
                "netamt"  : float(rec.NETAMT or 0),
                "rcvamt"  : float(rec.RCVAMT or 0),
                "remamt"  : remamt,
            })
        t.close()
        result.sort(key=lambda x: x["docdat"])
        return jsonify({"invoices": result})
    except Exception as e:
        return jsonify({"error":str(e)}), 500


@app.route('/browse-folder', methods=['GET'])
def browse_folder():
    """เปิด Windows folder picker dialog ผ่าน subprocess (thread-safe)"""
    try:
        import subprocess, sys, tempfile, json as _json

        initial = DBF_FOLDER if os.path.exists(DBF_FOLDER) else "Z:\\"
        # รัน tkinter dialog ใน subprocess แยกต่างหาก เพื่อหลีกเลี่ยงปัญหา
        # "main thread is not in main loop" ของ tkinter ใน Flask worker thread
        script = (
            "import tkinter as tk; from tkinter import filedialog; import json, sys;"
            f"root=tk.Tk(); root.withdraw(); root.lift(); root.attributes('-topmost',True);"
            f"f=filedialog.askdirectory(title='เลือกโฟลเดอร์ที่เก็บข้อมูล DBF',initialdir={repr(initial)});"
            "root.destroy();"
            "print(json.dumps({'path': f.replace('/','\\\\\\\\') if f else ''}))"
        )
        result = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True, text=True, timeout=120
        )
        if result.returncode != 0:
            raise Exception(result.stderr.strip() or "subprocess failed")

        out = result.stdout.strip()
        if not out:
            return jsonify({"success": False, "path": ""})
        data = _json.loads(out)
        path = data.get("path", "")
        if path:
            path = path.replace("/", "\\").rstrip("\\") + "\\"
            return jsonify({"success": True, "path": path})
        else:
            return jsonify({"success": False, "path": ""})
    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "path": "", "cancelled": True})
    except Exception as e:
        logger.error(f"browse-folder error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/list-drives', methods=['GET'])
def list_drives():
    """แสดงรายการ drive และ subfolder สำหรับ web-based folder browser"""
    try:
        req_path = request.args.get("path", "").strip()

        # ── ถ้าไม่ส่ง path มา → list drives ทั้งหมด ──
        if not req_path:
            import string
            drives = []
            for letter in string.ascii_uppercase:
                p = f"{letter}:\\"
                if os.path.exists(p):
                    drives.append({"name": p, "path": p, "type": "drive"})
            return jsonify({"path": "", "items": drives, "has_dbf": False})

        # ── normalize path ──
        req_path = req_path.rstrip("\\/") + "\\"
        if not os.path.isdir(req_path):
            return jsonify({"error": f"ไม่พบ: {req_path}"}), 404

        # ── list subfolder ──
        items = []
        try:
            with os.scandir(req_path) as it:
                for entry in it:
                    if entry.is_dir(follow_symlinks=False):
                        items.append({
                            "name": entry.name,
                            "path": os.path.join(req_path, entry.name) + "\\",
                            "type": "folder",
                        })
        except PermissionError:
            pass
        items.sort(key=lambda x: x["name"].lower())

        # ── ตรวจว่ามี DBF ไหม ──
        has_dbf = any(
            f.lower().endswith(".dbf")
            for f in os.listdir(req_path)
            if os.path.isfile(os.path.join(req_path, f))
        )
        gljnl_ok = os.path.exists(os.path.join(req_path, "GLJNL.DBF"))
        artrn_ok = os.path.exists(os.path.join(req_path, "ARTRN.DBF"))

        return jsonify({
            "path": req_path,
            "items": items,
            "has_dbf": has_dbf,
            "gljnl_ok": gljnl_ok,
            "artrn_ok": artrn_ok,
        })
    except Exception as e:
        logger.error(f"list-drives error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/validate-path', methods=['POST'])
def validate_path():
    """ตรวจสอบว่า path ที่ระบุมีไฟล์ DBF ของ Express หรือไม่"""
    try:
        data = request.json or {}
        path = data.get("path", "").strip().rstrip("\\/") + "\\"
        if not os.path.isdir(path):
            return jsonify({"valid": False, "error": "ไม่พบโฟลเดอร์นี้"})
        gljnl_ok = os.path.exists(os.path.join(path, "GLJNL.DBF"))
        artrn_ok = os.path.exists(os.path.join(path, "ARTRN.DBF"))
        has_dbf  = any(
            f.lower().endswith(".dbf")
            for f in os.listdir(path)
            if os.path.isfile(os.path.join(path, f))
        )
        return jsonify({
            "valid":    has_dbf,
            "gljnl_ok": gljnl_ok,
            "artrn_ok": artrn_ok,
            "has_dbf":  has_dbf,
            "path":     path,
        })
    except Exception as e:
        return jsonify({"valid": False, "error": str(e)}), 500


@app.route('/set-path', methods=['POST'])
def set_path():
    """เปลี่ยน DBF_FOLDER แบบ dynamic จากหน้าเว็บ"""
    global DBF_FOLDER
    try:
        data = request.json
        new_path = data.get("dbf_path", "").strip()
        if not new_path:
            return jsonify({"error": "dbf_path ว่างเปล่า"}), 400
        # normalize — เติม backslash ท้ายถ้าไม่มี
        new_path = new_path.rstrip("\\/") + "\\"
        DBF_FOLDER = new_path
        save_config(DBF_FOLDER)
        logger.info(f"[Agent] เปลี่ยน DBF_FOLDER → {DBF_FOLDER}")
        return jsonify({"success": True, "dbf_folder": DBF_FOLDER})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/status', methods=['GET'])
def status():
    dbf_ok = os.path.exists(os.path.join(DBF_FOLDER, "GLJNL.DBF"))
    ar_ok  = os.path.exists(os.path.join(DBF_FOLDER, "ARTRN.DBF"))
    return jsonify({"status":"online","dbf_folder":DBF_FOLDER,"dbf_ok":dbf_ok,"ar_ok":ar_ok,"version":"2.0.0"})

@app.route('/ping', methods=['GET'])
def ping():
    return jsonify({"pong": True})


# เพิ่ม route นำเข้า IN จาก Excel (/import-in, /inspect-in)
register_in_routes(app)


@app.route('/usage-history', methods=['GET'])
def usage_history():
    """ดึงประวัติการนำเข้าข้อมูล"""
    try:
        limit = int(request.args.get("limit", 20))
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                history = json.load(f)
        else:
            history = []
        return jsonify({"history": history[:limit]})
    except Exception as e:
        return jsonify({"history": [], "error": str(e)})


@app.route('/license-info', methods=['GET'])
def license_info():
    """ดึงข้อมูล license จาก license_checker"""
    try:
        from license_checker import get_license_info
        info = get_license_info()
        return jsonify(info)
    except Exception:
        # fallback — ถ้า license_checker ไม่มี get_license_info ให้ return ข้อมูลพื้นฐาน
        import uuid, platform
        machine_id = str(uuid.getnode())
        return jsonify({
            "plan": "รายเดือน",
            "machine_id": platform.node(),
            "expire_date": "30 เม.ย. 2569",
            "days_left": 24,
            "status": "active",
        })


@app.route('/payment-notify', methods=['POST'])
def payment_notify():
    """รับการแจ้งชำระเงินจากหน้าเว็บ"""
    try:
        data = request.json or {}
        record = {
            "id": int(datetime.now().timestamp() * 1000),
            "name": data.get("name", ""),
            "amount": data.get("amount", ""),
            "date": data.get("date", ""),
            "created_at": datetime.now().isoformat(),
            "status": "pending",
        }
        # บันทึกลงไฟล์
        payments = []
        if os.path.exists(PAYMENT_FILE):
            with open(PAYMENT_FILE, "r", encoding="utf-8") as f:
                payments = json.load(f)
        payments.insert(0, record)
        with open(PAYMENT_FILE, "w", encoding="utf-8") as f:
            json.dump(payments, f, ensure_ascii=False, indent=2)
        logger.info(f"[Payment] แจ้งชำระ: {record['name']} {record['amount']} บาท วันที่ {record['date']}")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Delete Routes ─────────────────────────────────────────────

@app.route('/delete/validate', methods=['POST'])
def delete_validate():
    """ตรวจสอบรายการก่อนลบ
    รับ: { doctype: "IN"|"JV"|"PV"|..., docnums: ["IN6904001", ...] }
    คืน: { results: [{docnum, cuscod, cusname, netamt, docstat, status, message}], ok, error }
    """
    try:
        data    = request.json or {}
        doctype = data.get("doctype", "IN").upper()
        docnums = [str(d).strip() for d in data.get("docnums", [])]
        if not docnums:
            return jsonify({"results": [], "ok": 0, "error": 0})

        results = []

        if doctype in ("IN", "RE"):
            artrn_path = os.path.join(DBF_FOLDER, "ARTRN.DBF")
            cust_names = load_cust_names()
            rectyp_filter = "3" if doctype == "IN" else "9"
            found = {}
            t = dbf.Table(artrn_path, codepage=ENCODING)
            t.open(mode=dbf.READ_ONLY)
            for rec in t:
                if dbf.is_deleted(rec): continue
                dn = str(rec.DOCNUM).strip()
                if dn in docnums and str(rec.RECTYP).strip() == rectyp_filter:
                    cuscod = str(rec.CUSCOD).strip()
                    found[dn] = {
                        "docnum":  dn,
                        "cuscod":  cuscod,
                        "cusname": cust_names.get(cuscod, ""),
                        "netamt":  float(rec.NETAMT or 0),
                        "docstat": str(rec.DOCSTAT).strip(),
                        "status":  "ok",
                        "message": "",
                    }
            t.close()
            for dn in docnums:
                if dn in found:
                    info = found[dn]
                    if doctype == "IN" and info["docstat"] == "Y":
                        info["status"]  = "warn"
                        info["message"] = "ชำระครบแล้ว"
                    results.append(info)
                else:
                    results.append({
                        "docnum": dn, "cuscod": "", "cusname": "",
                        "netamt": 0, "docstat": "",
                        "status": "error", "message": "ไม่พบเอกสาร",
                    })

        else:
            # GL types: JV, PV, RV, SV, UV, BW
            gljnl_path = os.path.join(DBF_FOLDER, "GLJNL.DBF")
            found = {}
            t = dbf.Table(gljnl_path, codepage=ENCODING)
            t.open(mode=dbf.READ_ONLY)
            for rec in t:
                if dbf.is_deleted(rec): continue
                vc = str(rec.VOUCHER).strip()
                if vc in docnums:
                    found[vc] = {
                        "docnum":  vc,
                        "cuscod":  "",
                        "cusname": str(getattr(rec, "DESCRP", "") or "").strip(),
                        "netamt":  0,
                        "docstat": str(getattr(rec, "DOCSTAT", "") or "").strip(),
                        "status":  "ok",
                        "message": "",
                    }
            t.close()
            for dn in docnums:
                if dn in found:
                    results.append(found[dn])
                else:
                    results.append({
                        "docnum": dn, "cuscod": "", "cusname": "",
                        "netamt": 0, "docstat": "",
                        "status": "error", "message": "ไม่พบเอกสาร",
                    })

        ok_count  = sum(1 for r in results if r["status"] in ("ok", "warn"))
        err_count = sum(1 for r in results if r["status"] == "error")
        logger.info(f"[Validate] {doctype} ok={ok_count} err={err_count}")
        return jsonify({"results": results, "ok": ok_count, "error": err_count})
    except Exception as e:
        logger.error(f"delete/validate error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/delete', methods=['POST'])
def do_delete():
    """ลบรายการออกจาก DBF (mark deleted)
    รับ: { doctype: "IN"|"JV"|"PV"|..., docnums: ["IN6904001", ...] }
    คืน: { success, error, details:[{docnum, status}] }
    IN  → ลบ ARTRN + STCRD + GLJNL + GLJNLIT
    GL  → ลบ GLJNL + GLJNLIT
    """
    try:
        data    = request.json or {}
        doctype = data.get("doctype", "IN").upper()
        docnums = set(str(d).strip() for d in data.get("docnums", []))
        logger.info(f"\n[Agent] ลบ {doctype} {len(docnums)} รายการ")

        deleted = set()

        if doctype == "RE":
            artrn_path   = os.path.join(DBF_FOLDER, "ARTRN.DBF")
            arrcpit_path = os.path.join(DBF_FOLDER, "ARRCPIT.DBF")
            arrcpcq_path = os.path.join(DBF_FOLDER, "ARRCPCQ.DBF")
            bktrn_path   = os.path.join(DBF_FOLDER, "BKTRN.DBF")
            gljnl_path   = os.path.join(DBF_FOLDER, "GLJNL.DBF")
            gljnlit_path = os.path.join(DBF_FOLDER, "GLJNLIT.DBF")

            # 1) ลบ ARTRN (RECTYP=9 = RE header)
            t = dbf.Table(artrn_path, codepage=ENCODING)
            t.open(mode=dbf.READ_WRITE)
            for rec in t:
                if dbf.is_deleted(rec): continue
                dn = str(rec.DOCNUM).strip()
                if dn in docnums and str(rec.RECTYP).strip() == "9":
                    dbf.delete(rec)
                    deleted.add(dn)
                    logger.info(f"  DEL ARTRN(RE) {dn}")
            t.close()

            # 2) ลบ ARRCPIT (matching items)
            if os.path.exists(arrcpit_path):
                t = dbf.Table(arrcpit_path, codepage=ENCODING)
                t.open(mode=dbf.READ_WRITE)
                for rec in t:
                    if dbf.is_deleted(rec): continue
                    if str(rec.RCPNUM).strip() in deleted:
                        dbf.delete(rec)
                t.close()

            # 3) ลบ ARRCPCQ (payment methods)
            if os.path.exists(arrcpcq_path):
                t = dbf.Table(arrcpcq_path, codepage=ENCODING)
                t.open(mode=dbf.READ_WRITE)
                for rec in t:
                    if dbf.is_deleted(rec): continue
                    if str(rec.RCPNUM).strip() in deleted:
                        dbf.delete(rec)
                t.close()

            # 4) ลบ BKTRN
            if os.path.exists(bktrn_path):
                t = dbf.Table(bktrn_path, codepage=ENCODING)
                t.open(mode=dbf.READ_WRITE)
                for rec in t:
                    if dbf.is_deleted(rec): continue
                    voucher = str(getattr(rec, "VOUCHER", "") or "").strip()
                    if voucher in deleted:
                        dbf.delete(rec)
                t.close()

            # 5) ลบ GLJNL + GLJNLIT
            if os.path.exists(gljnl_path):
                t = dbf.Table(gljnl_path, codepage=ENCODING)
                t.open(mode=dbf.READ_WRITE)
                for rec in t:
                    if dbf.is_deleted(rec): continue
                    if str(rec.VOUCHER).strip() in deleted:
                        dbf.delete(rec)
                t.close()
            if os.path.exists(gljnlit_path):
                t = dbf.Table(gljnlit_path, codepage=ENCODING)
                t.open(mode=dbf.READ_WRITE)
                for rec in t:
                    if dbf.is_deleted(rec): continue
                    if str(rec.VOUCHER).strip() in deleted:
                        dbf.delete(rec)
                t.close()

            log_usage(module="ลบรายการ — RE", count=len(deleted),
                      status="success" if len(deleted) == len(docnums) else "error")

        elif doctype == "IN":
            artrn_path   = os.path.join(DBF_FOLDER, "ARTRN.DBF")
            stcrd_path   = os.path.join(DBF_FOLDER, "STCRD.DBF")
            gljnl_path   = os.path.join(DBF_FOLDER, "GLJNL.DBF")
            gljnlit_path = os.path.join(DBF_FOLDER, "GLJNLIT.DBF")

            # 1) ลบ ARTRN (RECTYP=3 = IN header)
            t = dbf.Table(artrn_path, codepage=ENCODING)
            t.open(mode=dbf.READ_WRITE)
            for rec in t:
                if dbf.is_deleted(rec): continue
                dn = str(rec.DOCNUM).strip()
                if dn in docnums and str(rec.RECTYP).strip() == "3":
                    dbf.delete(rec)
                    deleted.add(dn)
                    logger.info(f"  DEL ARTRN {dn}")
            t.close()

            # 2) ลบ STCRD (line items)
            if os.path.exists(stcrd_path):
                t = dbf.Table(stcrd_path, codepage=ENCODING)
                t.open(mode=dbf.READ_WRITE)
                for rec in t:
                    if dbf.is_deleted(rec): continue
                    if str(rec.DOCNUM).strip() in deleted:
                        dbf.delete(rec)
                t.close()

            # 3) ลบ GLJNL header
            if os.path.exists(gljnl_path):
                t = dbf.Table(gljnl_path, codepage=ENCODING)
                t.open(mode=dbf.READ_WRITE)
                for rec in t:
                    if dbf.is_deleted(rec): continue
                    if str(rec.VOUCHER).strip() in deleted:
                        dbf.delete(rec)
                t.close()

            # 4) ลบ GLJNLIT line items
            if os.path.exists(gljnlit_path):
                t = dbf.Table(gljnlit_path, codepage=ENCODING)
                t.open(mode=dbf.READ_WRITE)
                for rec in t:
                    if dbf.is_deleted(rec): continue
                    if str(rec.VOUCHER).strip() in deleted:
                        dbf.delete(rec)
                t.close()

            log_usage(module="ลบรายการ — IN", count=len(deleted),
                      status="success" if len(deleted) == len(docnums) else "error")

        else:
            # GL types: JV, PV, RV, SV, UV, BW
            gljnl_path   = os.path.join(DBF_FOLDER, "GLJNL.DBF")
            gljnlit_path = os.path.join(DBF_FOLDER, "GLJNLIT.DBF")

            t = dbf.Table(gljnl_path, codepage=ENCODING)
            t.open(mode=dbf.READ_WRITE)
            for rec in t:
                if dbf.is_deleted(rec): continue
                vc = str(rec.VOUCHER).strip()
                if vc in docnums:
                    dbf.delete(rec)
                    deleted.add(vc)
                    logger.info(f"  DEL GLJNL {vc}")
            t.close()

            if os.path.exists(gljnlit_path):
                t = dbf.Table(gljnlit_path, codepage=ENCODING)
                t.open(mode=dbf.READ_WRITE)
                for rec in t:
                    if dbf.is_deleted(rec): continue
                    if str(rec.VOUCHER).strip() in deleted:
                        dbf.delete(rec)
                t.close()

            log_usage(module=f"ลบรายการ — {doctype}", count=len(deleted),
                      status="success" if len(deleted) == len(docnums) else "error")

        details = [
            {"docnum": dn, "status": "deleted" if dn in deleted else "not_found"}
            for dn in docnums
        ]
        ok  = len(deleted)
        err = len(docnums) - ok
        logger.info(f"[Delete] สำเร็จ {ok}, ไม่พบ {err}")
        return jsonify({"success": ok, "error": err, "details": details})
    except Exception as e:
        logger.error(f"delete error: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    logger.info("JW RPA Agent v2.0.0")
    logger.info(f"DBF Folder: {DBF_FOLDER}")
    logger.info(f"Port: {PORT}")
    auto_register()   # ลงทะเบียนเครื่องอัตโนมัติ
    app.run(host='127.0.0.1', port=PORT, debug=False)
